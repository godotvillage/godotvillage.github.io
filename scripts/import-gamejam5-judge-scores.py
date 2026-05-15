# -*- coding: utf-8 -*-
"""从三位评委 Excel 生成 gameJam5JudgeScores.ts 的 JSON 到 stdout"""
import json
import os
import re
import sys

try:
    import openpyxl
except ImportError:
    import subprocess
    subprocess.check_call([sys.executable, '-m', 'pip', 'install', 'openpyxl', '-q'])
    import openpyxl

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
ENTRIES_PATH = os.path.join(ROOT, 'src', 'data', 'gameJam5Entries.ts')
XLSX_DIR = os.environ.get('JAM5_SCORE_XLSX_DIR', r'd:\Download')

JUDGES = [
    {'id': 'bayun', 'name': '八云铀柚子', 'file': 'godot新手村第五届gamejam评分表 (八云铀柚子).xlsx'},
    {'id': 'qiyi', 'name': '祺奕祺可', 'file': 'Godot新手村第五届GameJam评分表 (祺奕祺可).xlsx'},
    {'id': 'lan', 'name': '斓', 'file': 'Godot新手村第五届GameJam评分表（斓）.xlsx'},
]


def parse_entries():
    content = open(ENTRIES_PATH, encoding='utf-8').read()
    entries = []
    pattern = re.compile(
        r"entryId:\s*'([^']+)',\s*work:\s*'((?:[^'\\]|\\.)*)',\s*author:\s*'((?:[^'\\]|\\.)*)'",
        re.MULTILINE,
    )
    for m in pattern.finditer(content):
        entries.append({
            'entryId': m.group(1),
            'work': m.group(2).replace("\\'", "'"),
            'author': m.group(3).replace("\\'", "'"),
        })
    if not entries:
        pattern2 = re.compile(
            r"entryId:\s*'([^']+)'[\s\S]*?work:\s*'([^']+)'[\s\S]*?author:\s*'([^']+)'",
            re.MULTILINE,
        )
        for m in pattern2.finditer(content):
            entries.append({'entryId': m.group(1), 'work': m.group(2), 'author': m.group(3)})
    return entries


def num(v):
    if v is None:
        return 0
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0


def str_val(v):
    if v is None:
        return ''
    return str(v).strip()


def main():
    entries = parse_entries()
    result = []
    for entry in entries:
        author = entry['author']
        reviews = []
        for judge in JUDGES:
            path = os.path.join(XLSX_DIR, judge['file'])
            wb = openpyxl.load_workbook(path, data_only=True)
            ws = wb[wb.sheetnames[0]]
            row_idx = None
            for r in range(2, ws.max_row + 1):
                if str_val(ws.cell(r, 1).value) == author:
                    row_idx = r
                    break
            if row_idx is None:
                print(f'WARN: missing score for {author} from {judge["name"]}', file=sys.stderr)
                continue
            r = row_idx
            reviews.append({
                'judgeId': judge['id'],
                'judgeName': judge['name'],
                'theme': num(ws.cell(r, 3).value),
                'themeNote': str_val(ws.cell(r, 4).value),
                'play': num(ws.cell(r, 5).value),
                'playNote': str_val(ws.cell(r, 6).value),
                'complete': num(ws.cell(r, 7).value),
                'completeNote': str_val(ws.cell(r, 8).value),
                'polish': num(ws.cell(r, 9).value),
                'polishNote': str_val(ws.cell(r, 10).value),
                'total': num(ws.cell(r, 11).value),
                'summary': str_val(ws.cell(r, 12).value),
            })
        if len(reviews) != 3:
            print(f'WARN: {author} has {len(reviews)} reviews', file=sys.stderr)
        avg = sum(x['total'] for x in reviews) / len(reviews) if reviews else 0
        result.append({
            'entryId': entry['entryId'],
            'author': author,
            'work': entry['work'],
            'reviews': reviews,
            'averageTotal': round(avg, 2),
            'displayScore': round(avg / 60 * 10, 1) if reviews else 0,
        })
    if hasattr(sys.stdout, 'reconfigure'):
        sys.stdout.reconfigure(encoding='utf-8')
    json.dump(result, sys.stdout, ensure_ascii=False)


if __name__ == '__main__':
    main()
